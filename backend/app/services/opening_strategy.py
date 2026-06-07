from math import ceil
from pathlib import Path
from typing import Any

import openpyxl
from pydantic import BaseModel, Field


FIRST_YEAR_SHEET = "\u7b2c\u4e00\u5e74"
CAPACITY_SHEET = "\u4ea7\u80fd\u8868"
SMART = "\u667a\u80fd"
AUTO = "\u81ea\u52a8"
TRADITIONAL = "\u4f20\u7edf"


class LineRule(BaseModel):
    name: str
    purchase_price: float
    install_quarters: int
    production_quarters: int = 1
    transfer_quarters: int = 0
    capacity_per_quarter: float
    production_fee_per_unit: float


class ProductRule(BaseModel):
    name: str
    material_cost: float
    sale_price: float
    qualification_quarter: int
    account_period_quarters: int = 4
    qualification_fee: float = 0

    @property
    def unit_margin(self) -> float:
        return self.sale_price - self.material_cost


class OpeningRules(BaseModel):
    initial_cash: float
    initial_equity: float
    loan_multiplier: float
    management_fee_per_quarter: float
    line_rules: dict[str, LineRule]
    products: dict[str, ProductRule]
    warnings: list[str] = Field(default_factory=list)

    @property
    def max_loan(self) -> float:
        return self.initial_equity * self.loan_multiplier


class QuarterlyCashFlow(BaseModel):
    quarter: int
    product: str | None
    capacity: float
    produced_units: float
    delivered_units: float
    beginning_cash: float
    cash_before_advertising: float
    advertising: float
    receivable_created: float
    collection: float
    material_cost: float
    production_fee: float
    line_purchase: float
    management_fee: float
    qualification_fee: float
    ending_cash: float


class OpeningPlanRow(BaseModel):
    plan_name: str
    line_config: dict[str, int]
    target_product: str
    total_capacity: float
    advertising_rank_requirement: str
    total_advertising: float
    net_profit: float
    cash_flow_ok: bool
    min_quarter_end_cash: float
    quarterly_cash_flow: list[QuarterlyCashFlow]
    risk_notes: list[str]


class OpeningPlanRequest(BaseModel):
    excel_path: str | None = None
    target_products: list[str] = Field(default_factory=lambda: ["P2"])
    max_lines: int = Field(default=8, ge=1, le=16)
    group_count: int | None = None
    market_demand: dict[str, float] = Field(default_factory=dict)
    competitor_capacities: list[float] = Field(default_factory=list)
    loan_usage_rate: float = Field(default=1.0, ge=0, le=1)
    include_mixed_lines: bool = True


class OpeningPlanResult(BaseModel):
    rules_summary: dict[str, Any]
    plans: list[OpeningPlanRow]
    warnings: list[str]


def _number(value: Any, default: float = 0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _normalize_line_type(raw: Any) -> str | None:
    text = str(raw or "")
    if AUTO in text:
        return "auto"
    if SMART in text:
        return "smart"
    if TRADITIONAL in text:
        return "traditional"
    normalized = text.strip().lower()
    if normalized in {"auto", "smart", "traditional"}:
        return normalized
    return None


def _line_defaults(line_type: str) -> tuple[float, int, int]:
    if line_type == "smart":
        return 200_000, 2, 0
    if line_type == "auto":
        return 100_000, 1, 0
    return 50_000, 0, 0


def _product_name(raw: Any) -> str | None:
    text = str(raw or "").strip().upper()
    if text in {"P1", "P2", "P3", "P4"}:
        return text
    return None


def _product_qualification_quarter(product: str) -> int:
    if product == "P1":
        return 1
    if product == "P2":
        return 3
    return 5


def _ad_quarters_for_year_one() -> set[int]:
    return {2, 3}


def load_opening_rules_from_excel(path: str | Path) -> OpeningRules:
    workbook = openpyxl.load_workbook(path, data_only=True)
    first = workbook[FIRST_YEAR_SHEET] if FIRST_YEAR_SHEET in workbook.sheetnames else workbook.active
    capacity_sheet = workbook[CAPACITY_SHEET] if CAPACITY_SHEET in workbook.sheetnames else None

    warnings: list[str] = []
    initial_equity = _number(first["N14"].value, _number(first["B2"].value, 0))
    loan_multiplier = _number(first["N15"].value, 3)
    management_fee = _number(first["N16"].value, _number(first["B11"].value, 15_000))

    purchase_prices = {
        "auto": _number(first["G30"].value, 100_000),
        "smart": _number(first["G32"].value, 200_000),
        "traditional": _number(first["G34"].value, 50_000),
    }

    line_rules: dict[str, LineRule] = {}
    if capacity_sheet is not None:
        line_rules.update(_load_line_rules(capacity_sheet, purchase_prices, warnings))
    if not line_rules:
        warnings.append("capacity table missing; using fallback line rules")
        line_rules = {
            line_type: LineRule(
                name=line_type,
                purchase_price=price,
                install_quarters=_line_defaults(line_type)[1],
                transfer_quarters=_line_defaults(line_type)[2],
                capacity_per_quarter={"auto": 46, "smart": 70, "traditional": 20}.get(line_type, 20),
                production_fee_per_unit={"auto": 400, "smart": 200, "traditional": 150}.get(line_type, 300),
            )
            for line_type, price in purchase_prices.items()
        }

    products = _load_product_rules(first, warnings)
    if not products:
        warnings.append("product price table missing; using fallback P1/P2 values")
        products = {
            "P1": ProductRule(name="P1", material_cost=1500, sale_price=3000, qualification_quarter=1),
            "P2": ProductRule(name="P2", material_cost=2500, sale_price=5000, qualification_quarter=3),
        }

    return OpeningRules(
        initial_cash=initial_equity,
        initial_equity=initial_equity,
        loan_multiplier=loan_multiplier,
        management_fee_per_quarter=management_fee,
        line_rules=line_rules,
        products=products,
        warnings=warnings,
    )


def _load_line_rules(
    capacity_sheet: Any,
    purchase_prices: dict[str, float],
    warnings: list[str],
) -> dict[str, LineRule]:
    variants: dict[str, list[LineRule]] = {}
    for type_row, capacity_row, fee_row in [(2, 11, 12), (15, 24, 25)]:
        for column in range(2, 10):
            line_type = _normalize_line_type(capacity_sheet.cell(type_row, column).value)
            if line_type is None:
                continue
            capacity = _number(capacity_sheet.cell(capacity_row, column).value, 0)
            if capacity <= 0:
                continue
            total_fee = _number(capacity_sheet.cell(fee_row, column).value, 0)
            default_price, install_quarters, transfer_quarters = _line_defaults(line_type)
            fee_per_unit = total_fee / capacity if total_fee > 0 else {"auto": 400, "smart": 200}.get(line_type, 150)
            variants.setdefault(line_type, []).append(
                LineRule(
                    name=line_type,
                    purchase_price=purchase_prices.get(line_type, default_price),
                    install_quarters=install_quarters,
                    production_quarters=1,
                    transfer_quarters=transfer_quarters,
                    capacity_per_quarter=capacity,
                    production_fee_per_unit=fee_per_unit,
                )
            )

    selected: dict[str, LineRule] = {}
    for line_type, rules in variants.items():
        selected[line_type] = max(rules, key=lambda rule: (rule.capacity_per_quarter, -rule.production_fee_per_unit))
        if len(rules) > 1:
            warnings.append(f"{line_type} has multiple capacity templates; selected highest quarterly capacity")
    return selected


def _load_product_rules(first: Any, warnings: list[str]) -> dict[str, ProductRule]:
    products: dict[str, ProductRule] = {}
    for row in range(30, 42):
        product = _product_name(first.cell(row, 9).value)
        if product is None:
            continue
        material_cost = _number(first.cell(row, 14).value, 0)
        sale_price = _number(first.cell(row, 15).value, 0)
        if sale_price <= 0:
            warnings.append(f"{product} sale price missing; skipped")
            continue
        products[product] = ProductRule(
            name=product,
            material_cost=material_cost,
            sale_price=sale_price,
            qualification_quarter=_product_qualification_quarter(product),
        )
    return products


def generate_opening_plans(
    rules: OpeningRules,
    target_products: list[str] | None = None,
    max_lines: int = 8,
    market_demand: dict[str, float] | None = None,
    competitor_capacities: list[float] | None = None,
    loan_usage_rate: float = 1.0,
    include_mixed_lines: bool = True,
) -> OpeningPlanResult:
    market_demand = market_demand or {}
    competitor_capacities = competitor_capacities or []
    target_products = target_products or ["P2"]
    max_lines = min(max_lines, 16)

    candidates = _candidate_line_configs(rules, max_lines, include_mixed_lines)
    plans: list[OpeningPlanRow] = []
    warnings = list(rules.warnings)

    for product_name in target_products:
        product = rules.products.get(product_name.upper())
        if product is None:
            warnings.append(f"{product_name} missing product rule; skipped")
            continue
        for line_config in candidates:
            plans.append(
                _simulate_candidate(
                    rules=rules,
                    line_config=line_config,
                    target_product=product,
                    market_demand=market_demand,
                    competitor_capacities=competitor_capacities,
                    loan_usage_rate=loan_usage_rate,
                )
            )

    plans.sort(key=lambda row: (row.cash_flow_ok, row.net_profit, row.min_quarter_end_cash), reverse=True)
    return OpeningPlanResult(rules_summary=_rules_summary(rules), plans=plans, warnings=warnings)


def _candidate_line_configs(
    rules: OpeningRules,
    max_lines: int,
    include_mixed_lines: bool,
) -> list[dict[str, int]]:
    configs: list[dict[str, int]] = []
    for line_type in rules.line_rules:
        for count in range(1, max_lines + 1):
            configs.append({line_type: count})

    if include_mixed_lines and "auto" in rules.line_rules and "smart" in rules.line_rules:
        for auto_count in range(1, max_lines):
            for smart_count in range(1, max_lines - auto_count + 1):
                configs.append({"auto": auto_count, "smart": smart_count})
    return configs


def _simulate_candidate(
    rules: OpeningRules,
    line_config: dict[str, int],
    target_product: ProductRule,
    market_demand: dict[str, float],
    competitor_capacities: list[float],
    loan_usage_rate: float,
) -> OpeningPlanRow:
    cash = rules.initial_cash + rules.max_loan * loan_usage_rate
    beginning_available_cash = cash
    total_ad = 0.0
    total_capacity = 0.0
    net_profit = 0.0
    min_cash = cash
    risk_notes: list[str] = []
    quarterly_rows: list[QuarterlyCashFlow] = []
    rank_requirement = "top 0"

    qualification_paid: set[str] = set()
    p1 = rules.products.get("P1")
    inventory: dict[str, float] = {}
    stock_arrivals: dict[int, dict[str, float]] = {}
    receivable_collections: dict[int, float] = {}
    pending_orders: dict[int, tuple[ProductRule, float]] = {}
    ad_quarters = _ad_quarters_for_year_one()

    for quarter in range(1, 9):
        beginning_cash = cash
        line_purchase = _line_purchase_in_quarter(rules, line_config, quarter)
        management = rules.management_fee_per_quarter
        product = target_product if quarter >= target_product.qualification_quarter else p1
        qualification_fee = 0.0
        if product is not None and product.name not in qualification_paid:
            qualification_fee = product.qualification_fee
            qualification_paid.add(product.name)

        active_capacity = _active_capacity(rules, line_config, quarter)
        capacity = active_capacity if product is not None else 0.0
        total_capacity += capacity

        for product_name, quantity in stock_arrivals.pop(quarter, {}).items():
            inventory[product_name] = inventory.get(product_name, 0.0) + quantity

        collection = receivable_collections.pop(quarter, 0.0)
        cash += collection

        advertising = 0.0
        current_order_units = 0.0
        if quarter in ad_quarters and product is not None:
            demand = _market_demand_for(product, capacity, market_demand, risk_notes)
            current_order_units = min(capacity, demand)
            advertising, rank_requirement = required_ad(current_order_units, demand, competitor_capacities)
            total_ad += advertising
            delivery_quarter = quarter + _production_cycle_for_quarter(rules, line_config, quarter)
            pending_orders[delivery_quarter] = (
                product,
                pending_orders.get(delivery_quarter, (product, 0.0))[1] + current_order_units,
            )

        avg_production_fee = _weighted_production_fee(rules, line_config, quarter)
        produced_units = current_order_units
        material_cost = produced_units * product.material_cost if product is not None else 0.0
        production_fee = produced_units * avg_production_fee
        if product is not None and produced_units > 0:
            stock_quarter = quarter + _production_cycle_for_quarter(rules, line_config, quarter)
            stock_arrivals.setdefault(stock_quarter, {})
            stock_arrivals[stock_quarter][product.name] = (
                stock_arrivals[stock_quarter].get(product.name, 0.0) + produced_units
            )

        order_product, ordered_units = pending_orders.pop(quarter, (None, 0.0))
        delivered_units = 0.0
        receivable_created = 0.0
        if order_product is not None and ordered_units > 0:
            available = inventory.get(order_product.name, 0.0)
            delivered_units = min(available, ordered_units)
            inventory[order_product.name] = available - delivered_units
            if delivered_units < ordered_units:
                risk_notes.append(f"Q{quarter} delivery shortfall for {order_product.name}")
            receivable_created = delivered_units * order_product.sale_price
            collection_quarter = quarter + order_product.account_period_quarters
            receivable_collections[collection_quarter] = (
                receivable_collections.get(collection_quarter, 0.0) + receivable_created
            )

        cash_before_ad = beginning_cash - management - line_purchase - qualification_fee
        if cash_before_ad < advertising:
            risk_notes.append(f"Q{quarter} cash before advertising is below required ad")

        cash = cash_before_ad - advertising - material_cost - production_fee + collection
        min_cash = min(min_cash, cash)
        net_profit += (
            receivable_created
            - material_cost
            - production_fee
            - management
            - advertising
            - qualification_fee
            - line_purchase
        )

        quarterly_rows.append(
            QuarterlyCashFlow(
                quarter=quarter,
                product=product.name if product is not None else None,
                capacity=capacity,
                produced_units=produced_units,
                delivered_units=delivered_units,
                beginning_cash=beginning_cash,
                cash_before_advertising=cash_before_ad,
                advertising=advertising,
                receivable_created=receivable_created,
                collection=collection,
                material_cost=material_cost,
                production_fee=production_fee,
                line_purchase=line_purchase,
                management_fee=management,
                qualification_fee=qualification_fee,
                ending_cash=cash,
            )
        )

    if min_cash < 0:
        risk_notes.insert(0, "cash flow becomes negative before delivery")
    if beginning_available_cash > rules.initial_cash:
        risk_notes.append("uses upfront loan capacity; financing timing should be refined against the official sheet")

    return OpeningPlanRow(
        plan_name=_plan_name(line_config, target_product.name),
        line_config=line_config,
        target_product=target_product.name,
        total_capacity=total_capacity,
        advertising_rank_requirement=rank_requirement,
        total_advertising=total_ad,
        net_profit=net_profit,
        cash_flow_ok=min_cash >= 0,
        min_quarter_end_cash=min_cash,
        quarterly_cash_flow=quarterly_rows,
        risk_notes=risk_notes,
    )


def required_ad(
    capacity: float,
    market_demand: float,
    competitor_capacities: list[float] | None = None,
) -> tuple[float, str]:
    if capacity <= 0:
        return 0.0, "top 0"
    competitor_capacities = competitor_capacities or []
    if competitor_capacities:
        rank = 1 + sum(1 for competitor_capacity in competitor_capacities if competitor_capacity > capacity)
    elif market_demand > 0:
        share = capacity / market_demand
        if share >= 0.3:
            rank = 1
        elif share >= 0.2:
            rank = 2
        elif share >= 0.1:
            rank = 3
        else:
            rank = 4
    else:
        rank = 4
    ad = capacity * 500 * (1 + 0.1 * rank)
    return ad, f"top {ceil(rank)}"


def _line_purchase_in_quarter(rules: OpeningRules, line_config: dict[str, int], quarter: int) -> float:
    if quarter != 1:
        return 0.0
    return sum(rules.line_rules[line_type].purchase_price * count for line_type, count in line_config.items())


def _active_capacity(rules: OpeningRules, line_config: dict[str, int], quarter: int) -> float:
    capacity = 0.0
    for line_type, count in line_config.items():
        line_rule = rules.line_rules[line_type]
        if quarter > line_rule.install_quarters:
            capacity += line_rule.capacity_per_quarter * count
    return capacity


def _production_cycle_for_quarter(rules: OpeningRules, line_config: dict[str, int], quarter: int) -> int:
    active_cycles = [
        rules.line_rules[line_type].production_quarters
        for line_type, count in line_config.items()
        if count > 0 and quarter > rules.line_rules[line_type].install_quarters
    ]
    return max(active_cycles) if active_cycles else 1


def _weighted_production_fee(rules: OpeningRules, line_config: dict[str, int], quarter: int) -> float:
    weighted_fee = 0.0
    total_capacity = 0.0
    for line_type, count in line_config.items():
        line_rule = rules.line_rules[line_type]
        if quarter <= line_rule.install_quarters:
            continue
        capacity = line_rule.capacity_per_quarter * count
        weighted_fee += capacity * line_rule.production_fee_per_unit
        total_capacity += capacity
    return weighted_fee / total_capacity if total_capacity else 0.0


def _market_demand_for(
    product: ProductRule | None,
    capacity: float,
    market_demand: dict[str, float],
    risk_notes: list[str],
) -> float:
    if product is None:
        return 0.0
    if product.name in market_demand:
        return market_demand[product.name]
    if capacity <= 0:
        return 0.0
    risk_notes.append(f"{product.name} market demand missing; using capacity-based placeholder demand")
    return capacity * 1.2


def _plan_name(line_config: dict[str, int], product_name: str) -> str:
    parts = [f"{line_type}x{count}" for line_type, count in sorted(line_config.items()) if count > 0]
    return f"{'+'.join(parts)}_{product_name}"


def _rules_summary(rules: OpeningRules) -> dict[str, Any]:
    return {
        "initial_cash": rules.initial_cash,
        "initial_equity": rules.initial_equity,
        "loan_multiplier": rules.loan_multiplier,
        "max_loan": rules.max_loan,
        "management_fee_per_quarter": rules.management_fee_per_quarter,
        "line_rules": {name: rule.model_dump() for name, rule in rules.line_rules.items()},
        "products": {name: rule.model_dump() | {"unit_margin": rule.unit_margin} for name, rule in rules.products.items()},
    }
