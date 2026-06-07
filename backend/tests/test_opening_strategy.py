from pathlib import Path

import openpyxl
from fastapi.testclient import TestClient

from app.main import app
from app.services.opening_strategy import (
    LineRule,
    OpeningRules,
    ProductRule,
    generate_opening_plans,
    load_opening_rules_from_excel,
)


def test_opening_strategy_generates_ranked_cash_checked_plans():
    rules = OpeningRules(
        initial_cash=880_000,
        initial_equity=880_000,
        loan_multiplier=3,
        management_fee_per_quarter=15_000,
        line_rules={
            "auto": LineRule(
                name="auto",
                purchase_price=100_000,
                install_quarters=1,
                production_quarters=1,
                transfer_quarters=0,
                capacity_per_quarter=46,
                production_fee_per_unit=400,
            ),
            "smart": LineRule(
                name="smart",
                purchase_price=200_000,
                install_quarters=2,
                production_quarters=1,
                transfer_quarters=0,
                capacity_per_quarter=70,
                production_fee_per_unit=200,
            ),
        },
        products={
            "P1": ProductRule(name="P1", material_cost=1500, sale_price=3000, qualification_quarter=1),
            "P2": ProductRule(name="P2", material_cost=2500, sale_price=5000, qualification_quarter=3),
        },
    )

    result = generate_opening_plans(rules, target_products=["P2"], max_lines=4)

    assert result.plans
    assert result.plans[0].line_config
    assert result.plans[0].target_product == "P2"
    assert result.plans[0].total_capacity > 0
    assert result.plans[0].total_advertising > 0
    assert result.plans[0].advertising_rank_requirement.startswith("top ")
    assert all(plan.min_quarter_end_cash >= 0 for plan in result.plans if plan.cash_flow_ok)


def test_opening_strategy_marks_negative_cash_flow_as_infeasible():
    rules = OpeningRules(
        initial_cash=100_000,
        initial_equity=100_000,
        loan_multiplier=0,
        management_fee_per_quarter=15_000,
        line_rules={
            "smart": LineRule(
                name="smart",
                purchase_price=200_000,
                install_quarters=2,
                production_quarters=1,
                transfer_quarters=0,
                capacity_per_quarter=70,
                production_fee_per_unit=200,
            ),
        },
        products={
            "P2": ProductRule(name="P2", material_cost=2500, sale_price=5000, qualification_quarter=3),
        },
    )

    result = generate_opening_plans(rules, target_products=["P2"], max_lines=1)

    assert result.plans[0].cash_flow_ok is False
    assert result.plans[0].min_quarter_end_cash < 0
    assert "cash flow" in result.plans[0].risk_notes[0]


def test_opening_strategy_respects_ad_production_delivery_and_account_period_timing():
    rules = OpeningRules(
        initial_cash=880_000,
        initial_equity=880_000,
        loan_multiplier=3,
        management_fee_per_quarter=15_000,
        line_rules={
            "auto": LineRule(
                name="auto",
                purchase_price=100_000,
                install_quarters=1,
                production_quarters=1,
                transfer_quarters=0,
                capacity_per_quarter=46,
                production_fee_per_unit=400,
            ),
        },
        products={
            "P1": ProductRule(name="P1", material_cost=1500, sale_price=3000, qualification_quarter=1),
            "P2": ProductRule(name="P2", material_cost=2500, sale_price=5000, qualification_quarter=3),
        },
    )

    result = generate_opening_plans(rules, target_products=["P2"], max_lines=1)
    plan = result.plans[0]

    q2 = plan.quarterly_cash_flow[1]
    q3 = plan.quarterly_cash_flow[2]
    q7 = plan.quarterly_cash_flow[6]

    assert q2.advertising > 0
    assert q2.produced_units == 46
    assert q2.delivered_units == 0
    assert q2.collection == 0

    assert q3.delivered_units == 46
    assert q3.receivable_created == 138_000
    assert q3.collection == 0

    assert q7.collection == 138_000


def test_load_opening_rules_from_excel_reads_key_cells(tmp_path: Path):
    workbook_path = tmp_path / "rules.xlsx"
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "\u7b2c\u4e00\u5e74"
    capacity = wb.create_sheet("\u4ea7\u80fd\u8868")
    wb.create_sheet("\u6750\u6599\u8868")
    wb.create_sheet("\u6fc0\u52b1\u8868")

    first["N14"] = 880_000
    first["N15"] = 3
    first["N16"] = 15_000
    first["G30"] = 100_000
    first["G32"] = 200_000
    first["G34"] = 50_000
    for row, product, material_cost, sale_price in [
        (34, "p1", 1500, 3000),
        (35, "p2", 2500, 5000),
        (36, "p3", 4000, 7000),
        (37, "P4", 6000, 13000),
    ]:
        first.cell(row=row, column=9, value=product)
        first.cell(row=row, column=14, value=material_cost)
        first.cell(row=row, column=15, value=sale_price)

    capacity["B2"] = "\u667a\u80fd\u7ebf"
    capacity["B3"] = 25
    capacity["B7"] = 0.9
    capacity["B8"] = 0.9
    capacity["B10"] = 1
    capacity["B11"] = 70
    capacity["B12"] = 14_000
    capacity["C2"] = "\u81ea\u52a8\u7ebf"
    capacity["C3"] = 20
    capacity["C5"] = 0.6
    capacity["C6"] = 0.6
    capacity["C7"] = 1
    capacity["C10"] = 1
    capacity["C11"] = 46
    capacity["C12"] = 18_400

    wb.save(workbook_path)

    rules = load_opening_rules_from_excel(workbook_path)

    assert rules.initial_cash == 880_000
    assert rules.loan_multiplier == 3
    assert rules.management_fee_per_quarter == 15_000
    assert rules.line_rules["smart"].capacity_per_quarter == 70
    assert rules.line_rules["auto"].purchase_price == 100_000
    assert rules.products["P2"].unit_margin == 2500


def test_opening_plan_api_accepts_excel_path(tmp_path: Path):
    workbook_path = tmp_path / "rules.xlsx"
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "\u7b2c\u4e00\u5e74"
    capacity = wb.create_sheet("\u4ea7\u80fd\u8868")
    wb.create_sheet("\u6750\u6599\u8868")
    first["N14"] = 880_000
    first["N15"] = 3
    first["N16"] = 15_000
    first["G30"] = 100_000
    first["G32"] = 200_000
    first["I35"] = "p2"
    first["N35"] = 2500
    first["O35"] = 5000
    capacity["B2"] = "\u81ea\u52a8\u7ebf"
    capacity["B11"] = 46
    capacity["B12"] = 18_400
    wb.save(workbook_path)

    with TestClient(app) as client:
        response = client.post(
            "/api/simulations/opening-plan",
            json={"excel_path": str(workbook_path), "target_products": ["P2"], "max_lines": 2},
        )

    assert response.status_code == 200
    data = response.json()
    assert data["plans"]
    assert data["plans"][0]["target_product"] == "P2"
